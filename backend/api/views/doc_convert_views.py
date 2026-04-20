import logging
import os
import subprocess
import tempfile
import uuid
from pathlib import Path

from django.http import HttpResponse
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

logger = logging.getLogger(__name__)

MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB

ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
    "application/vnd.oasis.opendocument.text",
}

ALLOWED_EXCEL_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

SOFFICE_CANDIDATES = [
    "libreoffice",
    "soffice",
    "/usr/local/bin/soffice",   # symlink created in Lambda container image
    "/usr/bin/libreoffice",
    "/usr/bin/soffice",
    "/opt/libreoffice/program/soffice",
    "/opt/instdir/program/soffice",  # Lambda layer path
    r"C:\Program Files\LibreOffice\program\soffice.exe",  # Windows default install
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
]


def _find_soffice() -> str | None:
    import shutil

    for candidate in SOFFICE_CANDIDATES:
        if shutil.which(candidate) or Path(candidate).exists():
            return candidate
    return None


def _convert_to_pdf(input_path: str, output_dir: str) -> str:
    """Run LibreOffice headless conversion and return the output PDF path."""
    soffice = _find_soffice()
    if soffice is None:
        raise RuntimeError(
            "LibreOffice is not installed. "
            "On Windows: winget install TheDocumentFoundation.LibreOffice\n"
            "On Linux:   apt-get install -y libreoffice\n"
            "On macOS:   brew install --cask libreoffice"
        )

    env = {
        **os.environ,
        "HOME": "/tmp",  # required in Lambda / read-only FS environments
    }

    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            output_dir,
            input_path,
        ],
        capture_output=True,
        text=True,
        timeout=120,
        env=env,
    )

    if result.returncode != 0:
        logger.error("LibreOffice stderr: %s", result.stderr)
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr[:500]}")

    stem = Path(input_path).stem
    pdf_path = Path(output_dir) / f"{stem}.pdf"
    if not pdf_path.exists():
        raise RuntimeError(
            f"Expected output PDF not found at {pdf_path}. "
            f"stdout={result.stdout!r} stderr={result.stderr!r}"
        )

    return str(pdf_path)


def _excel_color_to_reportlab(color_value: str):
    from reportlab.lib import colors

    if not color_value:
        return None
    if color_value.startswith("FF") and len(color_value) == 8:
        color_value = color_value[2:]
    if len(color_value) != 6:
        return None
    try:
        r = int(color_value[0:2], 16)
        g = int(color_value[2:4], 16)
        b = int(color_value[4:6], 16)
        return colors.Color(r / 255.0, g / 255.0, b / 255.0)
    except ValueError:
        return None


def _convert_excel_to_pdf_with_reportlab(input_path: str, output_path: str) -> None:
    from io import BytesIO
    from xml.sax.saxutils import escape

    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    with open(input_path, "rb") as fh:
        workbook_bytes = fh.read()

    wb = load_workbook(BytesIO(workbook_bytes), data_only=True)

    pagesize = landscape(A4)
    doc = SimpleDocTemplate(
        output_path,
        pagesize=pagesize,
        leftMargin=0.4 * inch,
        rightMargin=0.4 * inch,
        topMargin=0.4 * inch,
        bottomMargin=0.4 * inch,
    )

    story = []
    heading_style = ParagraphStyle(
        name="SheetHeading",
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        spaceAfter=8,
    )
    base_style = ParagraphStyle(
        name="Cell",
        fontName="Helvetica",
        fontSize=9,
        leading=11,
    )
    chunk_label_style = ParagraphStyle(
        name="ChunkLabel",
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        spaceBefore=6,
        spaceAfter=6,
    )

    try:
        for sheet_index, ws in enumerate(wb.worksheets):
            max_row = 0
            max_col = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        max_row = max(max_row, cell.row)
                        max_col = max(max_col, cell.column)

            if max_row == 0 or max_col == 0:
                continue

            data = []
            for r in range(1, max_row + 1):
                row_data = []
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    value = "" if cell.value is None else str(cell.value)
                    text = escape(value)
                    wrap = bool(cell.alignment and cell.alignment.wrap_text)
                    if wrap or "\n" in value:
                        row_data.append(Paragraph(text, base_style))
                    else:
                        row_data.append(text)
                data.append(row_data)

            col_widths = []
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                dim = ws.column_dimensions.get(col_letter)
                width = dim.width if dim and dim.width is not None else None
                if width is None:
                    max_len = 0
                    for r in range(1, max_row + 1):
                        value = ws.cell(row=r, column=c).value
                        max_len = max(max_len, len(str(value)) if value is not None else 0)
                    width = min(max(max_len, 6), 40)
                col_widths.append(width * 7)

            row_heights = []
            for r in range(1, max_row + 1):
                dim = ws.row_dimensions.get(r)
                height = dim.height if dim and dim.height is not None else 15
                row_heights.append(height)

            available_width = pagesize[0] - doc.leftMargin - doc.rightMargin

            def column_chunks():
                chunks = []
                start = 1
                width_acc = 0.0
                for idx, width in enumerate(col_widths, start=1):
                    if width_acc + width > available_width and width_acc > 0:
                        chunks.append((start, idx - 1))
                        start = idx
                        width_acc = 0.0
                    width_acc += width
                chunks.append((start, len(col_widths)))
                return chunks

            chunks = column_chunks()

            if sheet_index > 0:
                story.append(PageBreak())
            story.append(Paragraph(escape(ws.title), heading_style))

            for chunk_index, (start_col, end_col) in enumerate(chunks):
                chunk_data = [row[start_col - 1 : end_col] for row in data]
                chunk_col_widths = col_widths[start_col - 1 : end_col]
                chunk_width = sum(chunk_col_widths)
                if chunk_width > available_width and chunk_width > 0:
                    scale = available_width / chunk_width
                    chunk_col_widths = [w * scale for w in chunk_col_widths]

                if len(chunks) > 1:
                    start_letter = get_column_letter(start_col)
                    end_letter = get_column_letter(end_col)
                    label = f"Columns {start_letter}-{end_letter}"
                    story.append(Paragraph(label, chunk_label_style))

                table = Table(
                    chunk_data,
                    colWidths=chunk_col_widths,
                    rowHeights=row_heights,
                    repeatRows=1,
                )
                style_cmds = [
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]

                for merged_range in ws.merged_cells.ranges:
                    min_row = merged_range.min_row
                    max_row = merged_range.max_row
                    min_col = merged_range.min_col
                    max_col = merged_range.max_col
                    if (
                        min_row <= max_row
                        and min_col <= max_col
                        and min_col >= start_col
                        and max_col <= end_col
                    ):
                        style_cmds.append(
                            (
                                "SPAN",
                                (min_col - start_col, min_row - 1),
                                (max_col - start_col, max_row - 1),
                            )
                        )

                for r in range(1, max_row + 1):
                    for c in range(start_col, end_col + 1):
                        cell = ws.cell(row=r, column=c)
                        align = (cell.alignment.horizontal or "left").lower() if cell.alignment else "left"
                        valign = (cell.alignment.vertical or "top").lower() if cell.alignment else "top"

                        align_map = {
                            "left": "LEFT",
                            "center": "CENTER",
                            "right": "RIGHT",
                            "justify": "JUSTIFY",
                            "distributed": "LEFT",
                            "fill": "LEFT",
                        }
                        valign_map = {
                            "top": "TOP",
                            "center": "MIDDLE",
                            "bottom": "BOTTOM",
                            "justify": "MIDDLE",
                            "distributed": "MIDDLE",
                        }

                        style_cmds.append(
                            (
                                "ALIGN",
                                (c - start_col, r - 1),
                                (c - start_col, r - 1),
                                align_map.get(align, "LEFT"),
                            )
                        )
                        style_cmds.append(
                            (
                                "VALIGN",
                                (c - start_col, r - 1),
                                (c - start_col, r - 1),
                                valign_map.get(valign, "TOP"),
                            )
                        )

                        if cell.font:
                            if cell.font.bold:
                                style_cmds.append(
                                    (
                                        "FONTNAME",
                                        (c - start_col, r - 1),
                                        (c - start_col, r - 1),
                                        "Helvetica-Bold",
                                    )
                                )
                            if cell.font.sz:
                                style_cmds.append(
                                    (
                                        "FONTSIZE",
                                        (c - start_col, r - 1),
                                        (c - start_col, r - 1),
                                        float(cell.font.sz),
                                    )
                                )

                        fill = cell.fill
                        if fill and fill.fill_type == "solid" and fill.fgColor is not None:
                            bg_color = _excel_color_to_reportlab(fill.fgColor.rgb)
                            if bg_color:
                                style_cmds.append(
                                    (
                                        "BACKGROUND",
                                        (c - start_col, r - 1),
                                        (c - start_col, r - 1),
                                        bg_color,
                                    )
                                )

                table.setStyle(TableStyle(style_cmds))
                story.append(table)
                story.append(Spacer(1, 12))

        if not story:
            raise RuntimeError("No data found in the workbook.")

        doc.build(story)
    finally:
        wb.close()


@api_view(["POST"])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def word_to_pdf_view(request):
    doc_file = request.FILES.get("file")
    if doc_file is None:
        return Response({"error": "No file uploaded. Send a DOCX file as 'file'."}, status=400)

    content_type = (doc_file.content_type or "").split(";")[0].strip().lower()
    if content_type not in ALLOWED_MIME_TYPES:
        # Also accept by extension for browsers that send generic MIME types
        name_lower = (doc_file.name or "").lower()
        if not (name_lower.endswith(".docx") or name_lower.endswith(".doc") or name_lower.endswith(".odt")):
            return Response(
                {
                    "error": (
                        f"Unsupported file type '{content_type}'. "
                        "Please upload a .docx, .doc, or .odt file."
                    )
                },
                status=415,
            )

    if doc_file.size > MAX_UPLOAD_BYTES:
        return Response(
            {"error": f"File too large. Maximum allowed size is {MAX_UPLOAD_BYTES // 1024 // 1024} MB."},
            status=413,
        )

    with tempfile.TemporaryDirectory(prefix="torensa_docconv_") as tmpdir:
        # Use a safe unique input filename to avoid collisions
        name_lower_ext = (doc_file.name or "").lower()
        if name_lower_ext.endswith(".docx"):
            ext = ".docx"
        elif name_lower_ext.endswith(".odt"):
            ext = ".odt"
        else:
            ext = ".doc"
        input_filename = f"{uuid.uuid4().hex}{ext}"
        input_path = os.path.join(tmpdir, input_filename)

        with open(input_path, "wb") as fh:
            for chunk in doc_file.chunks():
                fh.write(chunk)

        try:
            pdf_path = _convert_to_pdf(input_path, tmpdir)
        except RuntimeError as exc:
            logger.exception("Word-to-PDF conversion error")
            return Response({"error": str(exc)}, status=500)
        except subprocess.TimeoutExpired:
            return Response({"error": "Conversion timed out. Please try a smaller file."}, status=500)

        with open(pdf_path, "rb") as fh:
            pdf_bytes = fh.read()

    original_stem = Path(doc_file.name or "document").stem
    output_name = f"{original_stem}.pdf"

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{output_name}"'
    response["Content-Length"] = str(len(pdf_bytes))
    return response


@api_view(["POST"])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def excel_to_pdf_view(request):
    excel_file = request.FILES.get("file")
    if excel_file is None:
        return Response(
            {"error": "No file uploaded. Send an Excel file as 'file'."},
            status=400,
        )

    content_type = (excel_file.content_type or "").split(";")[0].strip().lower()
    if content_type not in ALLOWED_EXCEL_MIME_TYPES:
        name_lower = (excel_file.name or "").lower()
        if not (
            name_lower.endswith(".xlsx")
            or name_lower.endswith(".xls")
            or name_lower.endswith(".xlsm")
        ):
            return Response(
                {
                    "error": (
                        f"Unsupported file type '{content_type}'. "
                        "Please upload a .xlsx, .xls, or .xlsm file."
                    )
                },
                status=415,
            )

    if excel_file.size > MAX_UPLOAD_BYTES:
        return Response(
            {
                "error": (
                    "File too large. Maximum allowed size is "
                    f"{MAX_UPLOAD_BYTES // 1024 // 1024} MB."
                )
            },
            status=413,
        )

    with tempfile.TemporaryDirectory(prefix="torensa_docconv_") as tmpdir:
        name_lower = (excel_file.name or "").lower()
        if name_lower.endswith(".xlsm"):
            ext = ".xlsm"
        elif name_lower.endswith(".xls"):
            ext = ".xls"
        else:
            ext = ".xlsx"

        input_filename = f"{uuid.uuid4().hex}{ext}"
        input_path = os.path.join(tmpdir, input_filename)

        with open(input_path, "wb") as fh:
            for chunk in excel_file.chunks():
                fh.write(chunk)

        try:
            if ext == ".xls":
                pdf_path = _convert_to_pdf(input_path, tmpdir)
            else:
                pdf_path = os.path.join(tmpdir, f"{Path(input_path).stem}.pdf")
                _convert_excel_to_pdf_with_reportlab(input_path, pdf_path)
        except RuntimeError as exc:
            logger.exception("Excel-to-PDF conversion error")
            return Response({"error": str(exc)}, status=500)
        except subprocess.TimeoutExpired:
            return Response(
                {"error": "Conversion timed out. Please try a smaller file."},
                status=500,
            )

        with open(pdf_path, "rb") as fh:
            pdf_bytes = fh.read()

    original_stem = Path(excel_file.name or "spreadsheet").stem
    output_name = f"{original_stem}.pdf"

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{output_name}"'
    response["Content-Length"] = str(len(pdf_bytes))
    return response


def _convert_to_docx(input_path: str, output_dir: str) -> str:
    """Run LibreOffice headless PDF→DOCX conversion and return the output path."""
    soffice = _find_soffice()
    if soffice is None:
        raise RuntimeError(
            "LibreOffice is not installed. "
            "On Windows: winget install TheDocumentFoundation.LibreOffice\n"
            "On Linux:   apt-get install -y libreoffice\n"
            "On macOS:   brew install --cask libreoffice"
        )

    env = {
        **os.environ,
        "HOME": "/tmp",
        "SAL_USE_VCLPLUGIN": "svp",
    }

    result = subprocess.run(
        [
            soffice,
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            "--infilter=writer_pdf_import",
            "--convert-to",
            "docx",
            "--outdir",
            output_dir,
            input_path,
        ],
        capture_output=True,
        text=True,
        timeout=120,
        env=env,
    )

    if result.returncode != 0:
        logger.error("LibreOffice stderr: %s", result.stderr)
        raise RuntimeError(f"LibreOffice conversion failed: {result.stderr[:500]}")

    stem = Path(input_path).stem
    docx_path = Path(output_dir) / f"{stem}.docx"
    if not docx_path.exists():
        raise RuntimeError(
            f"Expected output DOCX not found at {docx_path}. "
            f"stdout={result.stdout!r} stderr={result.stderr!r}"
        )

    return str(docx_path)


@api_view(["POST"])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def pdf_to_word_view(request):
    pdf_file = request.FILES.get("file")
    if pdf_file is None:
        return Response({"error": "No file uploaded. Send a PDF file as 'file'."}, status=400)

    name_lower = (pdf_file.name or "").lower()
    content_type = (pdf_file.content_type or "").split(";")[0].strip().lower()
    if content_type != "application/pdf" and not name_lower.endswith(".pdf"):
        return Response(
            {"error": f"Unsupported file type '{content_type}'. Please upload a .pdf file."},
            status=415,
        )

    if pdf_file.size > MAX_UPLOAD_BYTES:
        return Response(
            {"error": f"File too large. Maximum allowed size is {MAX_UPLOAD_BYTES // 1024 // 1024} MB."},
            status=413,
        )

    with tempfile.TemporaryDirectory(prefix="torensa_docconv_") as tmpdir:
        input_filename = f"{uuid.uuid4().hex}.pdf"
        input_path = os.path.join(tmpdir, input_filename)

        with open(input_path, "wb") as fh:
            for chunk in pdf_file.chunks():
                fh.write(chunk)

        try:
            docx_path = _convert_to_docx(input_path, tmpdir)
        except RuntimeError as exc:
            logger.exception("PDF-to-Word conversion error")
            return Response({"error": str(exc)}, status=500)
        except subprocess.TimeoutExpired:
            return Response({"error": "Conversion timed out. Please try a smaller file."}, status=500)

        with open(docx_path, "rb") as fh:
            docx_bytes = fh.read()

    original_stem = Path(pdf_file.name or "document").stem
    output_name = f"{original_stem}.docx"

    response = HttpResponse(
        docx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    response["Content-Disposition"] = f'attachment; filename="{output_name}"'
    response["Content-Length"] = str(len(docx_bytes))
    return response


@api_view(["POST"])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def pdf_extract_text_view(request):
    pdf_file = request.FILES.get("file")
    if pdf_file is None:
        return Response({"error": "No file uploaded. Send a PDF file as 'file'."}, status=400)

    name_lower = (pdf_file.name or "").lower()
    content_type = (pdf_file.content_type or "").split(";")[0].strip().lower()
    if content_type != "application/pdf" and not name_lower.endswith(".pdf"):
        return Response(
            {"error": f"Unsupported file type '{content_type}'. Please upload a .pdf file."},
            status=415,
        )

    if pdf_file.size > MAX_UPLOAD_BYTES:
        return Response(
            {"error": f"File too large. Maximum allowed size is {MAX_UPLOAD_BYTES // 1024 // 1024} MB."},
            status=413,
        )

    with tempfile.TemporaryDirectory(prefix="torensa_pdfocr_") as tmpdir:
        input_filename = f"{uuid.uuid4().hex}.pdf"
        input_path = os.path.join(tmpdir, input_filename)

        with open(input_path, "wb") as fh:
            for chunk in pdf_file.chunks():
                fh.write(chunk)

        try:
            text = _extract_text_from_pdf(input_path)
        except RuntimeError as exc:
            logger.exception("PDF text extraction error")
            return Response({"error": str(exc)}, status=500)

    return Response({"text": text})


def _extract_text_from_pdf(input_path: str) -> str:
    """Extract text from a PDF, falling back to OCR for scanned pages."""
    import fitz  # PyMuPDF

    doc = fitz.open(input_path)
    pages_text = []

    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text().strip()

            if len(text) < 20:
                # Likely a scanned page — try OCR
                ocr_text = _ocr_page(page)
                if ocr_text:
                    text = ocr_text

            pages_text.append(text)
    finally:
        doc.close()

    return "\n\n".join(pages_text)


def _ocr_page(page) -> str:
    """Render a PDF page to an image and run Tesseract OCR on it."""
    try:
        import pytesseract
        from PIL import Image
        from io import BytesIO

        # Render at 300 DPI for good OCR quality
        pix = page.get_pixmap(dpi=300)
        img = Image.open(BytesIO(pix.tobytes("png")))
        return pytesseract.image_to_string(img).strip()
    except Exception as exc:
        logger.warning("OCR failed for page %s: %s", page.number, exc)
        return ""
